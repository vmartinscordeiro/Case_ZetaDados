"""
Data Catalog Generator
======================
Generates a formatted Excel data dictionary (data_dictionary.xlsx) that documents:

  - All tables across Bronze / Silver / Gold layers
  - Column metadata: type, nullable, description, validations, example values
  - Quality profile: null rate and sample values read from actual Silver CSVs

Output
------
    docs/data_dictionary.xlsx   (3 sheets: Visão Geral · Colunas · Regras de Qualidade)

Usage
-----
    python docs/generate_catalog.py
"""

import sys
from pathlib import Path

import pandas as pd

# Make project root importable when running from docs/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config.settings import SILVER_DIR, GOLD_DIR

# ─────────────────────────────────────────────────────────────────────────────
# CATALOG DEFINITION
# Each entry = one column across any layer.
# ─────────────────────────────────────────────────────────────────────────────

CATALOG: list[dict] = [

    # ── BRONZE ── orders ─────────────────────────────────────────────────────
    dict(layer="Bronze", table="orders", column="order_id",
         type="str",   nullable=False, pk=True,
         description="Identificador único do pedido.",
         validation="Não pode ser nulo. Deve ser único por pedido.",
         origin="Fonte", example="O000001"),
    dict(layer="Bronze", table="orders", column="customer_id",
         type="str",   nullable=False, pk=False,
         description="Referência ao cliente que realizou o pedido.",
         validation="Não pode ser nulo.",
         origin="Fonte", example="C00083"),
    dict(layer="Bronze", table="orders", column="order_ts",
         type="str",   nullable=True,  pk=False,
         description="Timestamp em que o pedido foi criado (formato variável na origem).",
         validation="Pode chegar nulo ou em formatos DD/MM/YYYY HH:MM ou YYYY-MM-DD HH:MM:SS.",
         origin="Fonte", example="01/12/2025 05:57"),
    dict(layer="Bronze", table="orders", column="status",
         type="str",   nullable=True,  pk=False,
         description="Status atual do pedido na origem.",
         validation="Aceita qualquer string; normalizado no Silver.",
         origin="Fonte", example="paid"),
    dict(layer="Bronze", table="orders", column="payment_method",
         type="str",   nullable=True,  pk=False,
         description="Forma de pagamento utilizada.",
         validation="Aceita qualquer string; normalizado no Silver.",
         origin="Fonte", example="debit_card"),
    dict(layer="Bronze", table="orders", column="total_amount",
         type="object", nullable=True, pk=False,
         description="Valor bruto do pedido na moeda indicada.",
         validation="Numérico positivo; coerção para float aplicada no Silver.",
         origin="Fonte", example="48.82"),
    dict(layer="Bronze", table="orders", column="currency",
         type="str",   nullable=True,  pk=False,
         description="Moeda do pedido.",
         validation="Normalizado para maiúsculas no Silver (ex.: BRL).",
         origin="Fonte", example="BRL"),

    # ── BRONZE ── order_items ────────────────────────────────────────────────
    dict(layer="Bronze", table="order_items", column="order_id",
         type="str",   nullable=False, pk=True,
         description="Referência ao pedido pai.",
         validation="FK para orders.order_id.",
         origin="Fonte", example="O000001"),
    dict(layer="Bronze", table="order_items", column="product_id",
         type="str",   nullable=False, pk=True,
         description="Referência ao produto comprado.",
         validation="FK para products.product_id.",
         origin="Fonte", example="P00029"),
    dict(layer="Bronze", table="order_items", column="quantity",
         type="object", nullable=True, pk=False,
         description="Quantidade de unidades compradas.",
         validation="Inteiro positivo; coerção no Silver.",
         origin="Fonte", example="2"),
    dict(layer="Bronze", table="order_items", column="unit_price",
         type="object", nullable=True, pk=False,
         description="Preço unitário do produto em BRL.",
         validation="Numérico positivo.",
         origin="Fonte", example="567.16"),
    dict(layer="Bronze", table="order_items", column="discount_amount",
         type="object", nullable=True, pk=False,
         description="Desconto total aplicado ao item em BRL.",
         validation="Numérico ≥ 0; nulos tratados como 0 no Silver.",
         origin="Fonte", example="21.74"),

    # ── BRONZE ── shipments ──────────────────────────────────────────────────
    dict(layer="Bronze", table="shipments", column="order_id",
         type="str",   nullable=False, pk=True,
         description="Referência ao pedido enviado.",
         validation="FK para orders.order_id.",
         origin="Fonte", example="O000001"),
    dict(layer="Bronze", table="shipments", column="carrier",
         type="str",   nullable=True,  pk=False,
         description="Nome da transportadora.",
         validation="Pode chegar com espaços extras; normalizado no Silver.",
         origin="Fonte", example="CORREIOS"),
    dict(layer="Bronze", table="shipments", column="shipping_cost",
         type="object", nullable=True, pk=False,
         description="Custo do frete em BRL.",
         validation="Pode usar vírgula decimal (ex.: '0,00'); convertido no Silver.",
         origin="Fonte", example='"0,00"'),
    dict(layer="Bronze", table="shipments", column="shipped_ts",
         type="str",   nullable=True,  pk=False,
         description="Timestamp de despacho do pedido.",
         validation="Formatos mistos; parseado com dayfirst=True no Silver.",
         origin="Fonte", example="02/12/2025 07:22"),
    dict(layer="Bronze", table="shipments", column="delivered_ts",
         type="str",   nullable=True,  pk=False,
         description="Timestamp de entrega ao cliente.",
         validation="Pode conter 'N/A' ou estar vazio; convertido para NaT no Silver.",
         origin="Fonte", example="2025-12-07 00:08:00"),
    dict(layer="Bronze", table="shipments", column="delivery_status",
         type="str",   nullable=True,  pk=False,
         description="Status final da entrega.",
         validation="Normalizado para lowercase no Silver.",
         origin="Fonte", example="returned"),

    # ── BRONZE ── customers ──────────────────────────────────────────────────
    dict(layer="Bronze", table="customers", column="customer_id",
         type="str",   nullable=False, pk=True,
         description="Identificador único do cliente.",
         validation="Não pode ser nulo.",
         origin="Fonte", example="C00001"),
    dict(layer="Bronze", table="customers", column="state",
         type="str",   nullable=True,  pk=False,
         description="UF (estado) do cliente.",
         validation="Pode chegar em casing inconsistente (ex.: 'Rj'); padronizado para maiúsculas.",
         origin="Fonte", example="SP"),
    dict(layer="Bronze", table="customers", column="city",
         type="str",   nullable=True,  pk=False,
         description="Cidade do cliente.",
         validation="Padronizado para Title Case no Silver.",
         origin="Fonte", example="São Paulo"),
    dict(layer="Bronze", table="customers", column="created_ts",
         type="str",   nullable=True,  pk=False,
         description="Timestamp de cadastro do cliente.",
         validation="Formatos mistos; parseado com dayfirst=True.",
         origin="Fonte", example="2023-04-13 00:00:00"),

    # ── BRONZE ── products ───────────────────────────────────────────────────
    dict(layer="Bronze", table="products", column="product_id",
         type="str",   nullable=False, pk=True,
         description="Identificador único do produto.",
         validation="Não pode ser nulo.",
         origin="Fonte", example="P00001"),
    dict(layer="Bronze", table="products", column="category",
         type="str",   nullable=True,  pk=False,
         description="Categoria do produto.",
         validation="Texto livre; espaços extras removidos no Silver.",
         origin="Fonte", example="Eletrônicos"),
    dict(layer="Bronze", table="products", column="brand",
         type="str",   nullable=True,  pk=False,
         description="Marca do produto.",
         validation="Pode ser nulo.",
         origin="Fonte", example="Acme"),
    dict(layer="Bronze", table="products", column="created_ts",
         type="str",   nullable=True,  pk=False,
         description="Timestamp de cadastro do produto.",
         validation="Formatos mistos; parseado com dayfirst=True.",
         origin="Fonte", example="18/03/2024 00:00"),

    # ── SILVER ── orders (colunas adicionadas/transformadas) ─────────────────
    dict(layer="Silver", table="orders", column="order_id",
         type="str",      nullable=False, pk=True,
         description="Identificador único do pedido.",
         validation="Deduplicado por order_id (keep='last').",
         origin="Bronze", example="O000001"),
    dict(layer="Silver", table="orders", column="customer_id",
         type="str",      nullable=False, pk=False,
         description="Referência ao cliente.",
         validation="Não nulo.",
         origin="Bronze", example="C00083"),
    dict(layer="Silver", table="orders", column="order_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp do pedido convertido para datetime.",
         validation="Nulos mantidos como NaT se o formato for inválido.",
         origin="Bronze", example="2025-12-01 05:57:00"),
    dict(layer="Silver", table="orders", column="order_date",
         type="datetime", nullable=True,  pk=False,
         description="Data do pedido (sem horário), derivada de order_ts.",
         validation="Derivada: order_ts.dt.normalize().",
         origin="Derivada", example="2025-12-01"),
    dict(layer="Silver", table="orders", column="status",
         type="str",      nullable=True,  pk=False,
         description="Status normalizado para lowercase.",
         validation="Valores esperados: paid, created, cancelled, returned.",
         origin="Bronze", example="paid"),
    dict(layer="Silver", table="orders", column="payment_method",
         type="str",      nullable=True,  pk=False,
         description="Forma de pagamento normalizada para lowercase.",
         validation="Valores esperados: pix, credit_card, debit_card.",
         origin="Bronze", example="debit_card"),
    dict(layer="Silver", table="orders", column="total_amount",
         type="float",    nullable=True,  pk=False,
         description="Valor bruto do pedido em BRL.",
         validation="float ≥ 0.",
         origin="Bronze", example="48.82"),
    dict(layer="Silver", table="orders", column="currency",
         type="str",      nullable=True,  pk=False,
         description="Moeda normalizada para maiúsculas.",
         validation="Padrão: BRL.",
         origin="Bronze", example="BRL"),

    # ── SILVER ── order_items ────────────────────────────────────────────────
    dict(layer="Silver", table="order_items", column="order_id",
         type="str",   nullable=False, pk=True,
         description="Referência ao pedido pai.",
         validation="FK para silver.orders.",
         origin="Bronze", example="O000001"),
    dict(layer="Silver", table="order_items", column="product_id",
         type="str",   nullable=False, pk=True,
         description="Referência ao produto.",
         validation="FK para silver.products.",
         origin="Bronze", example="P00029"),
    dict(layer="Silver", table="order_items", column="quantity",
         type="int",   nullable=False, pk=False,
         description="Quantidade de unidades.",
         validation="int ≥ 0; nulos preenchidos com 0.",
         origin="Bronze", example="2"),
    dict(layer="Silver", table="order_items", column="unit_price",
         type="float", nullable=True,  pk=False,
         description="Preço unitário em BRL.",
         validation="float ≥ 0.",
         origin="Bronze", example="567.16"),
    dict(layer="Silver", table="order_items", column="discount_amount",
         type="float", nullable=False, pk=False,
         description="Desconto total do item em BRL.",
         validation="float ≥ 0; nulos preenchidos com 0.",
         origin="Bronze", example="21.74"),
    dict(layer="Silver", table="order_items", column="item_net_amount",
         type="float", nullable=True,  pk=False,
         description="Receita líquida do item: (quantity × unit_price) − discount_amount.",
         validation="Derivada no Silver.",
         origin="Derivada", example="1113.32"),

    # ── SILVER ── shipments ──────────────────────────────────────────────────
    dict(layer="Silver", table="shipments", column="order_id",
         type="str",      nullable=False, pk=True,
         description="Referência ao pedido.",
         validation="Deduplicado por order_id.",
         origin="Bronze", example="O000001"),
    dict(layer="Silver", table="shipments", column="carrier",
         type="str",      nullable=True,  pk=False,
         description="Transportadora com espaços removidos.",
         validation="Texto normalizado.",
         origin="Bronze", example="CORREIOS"),
    dict(layer="Silver", table="shipments", column="shipping_cost",
         type="float",    nullable=True,  pk=False,
         description="Custo do frete em BRL (vírgula decimal corrigida).",
         validation="float ≥ 0.",
         origin="Bronze", example="0.00"),
    dict(layer="Silver", table="shipments", column="shipped_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp de despacho convertido para datetime.",
         validation="NaT se inválido.",
         origin="Bronze", example="2025-12-02 07:22:00"),
    dict(layer="Silver", table="shipments", column="delivered_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp de entrega convertido para datetime.",
         validation="'N/A' e variantes convertidos para NaT.",
         origin="Bronze", example="2025-12-07 00:08:00"),
    dict(layer="Silver", table="shipments", column="delivery_status",
         type="str",      nullable=True,  pk=False,
         description="Status de entrega normalizado para lowercase.",
         validation="Valores esperados: delivered, returned, in_transit.",
         origin="Bronze", example="returned"),

    # ── SILVER ── customers ──────────────────────────────────────────────────
    dict(layer="Silver", table="customers", column="customer_id",
         type="str",      nullable=False, pk=True,
         description="Identificador único do cliente.",
         validation="Deduplicado.",
         origin="Bronze", example="C00001"),
    dict(layer="Silver", table="customers", column="state",
         type="str",      nullable=True,  pk=False,
         description="UF normalizada para maiúsculas (ex.: 'Rj' → 'RJ').",
         validation="2 caracteres maiúsculos.",
         origin="Bronze", example="RJ"),
    dict(layer="Silver", table="customers", column="city",
         type="str",      nullable=True,  pk=False,
         description="Cidade em Title Case.",
         validation="Texto normalizado.",
         origin="Bronze", example="Rio De Janeiro"),
    dict(layer="Silver", table="customers", column="created_ts",
         type="datetime", nullable=True,  pk=False,
         description="Data de cadastro do cliente.",
         validation="NaT se inválido.",
         origin="Bronze", example="2023-04-13 00:00:00"),

    # ── SILVER ── products ───────────────────────────────────────────────────
    dict(layer="Silver", table="products", column="product_id",
         type="str",      nullable=False, pk=True,
         description="Identificador único do produto.",
         validation="Deduplicado.",
         origin="Bronze", example="P00001"),
    dict(layer="Silver", table="products", column="category",
         type="str",      nullable=True,  pk=False,
         description="Categoria do produto (espaços removidos).",
         validation="Texto normalizado.",
         origin="Bronze", example="Eletrônicos"),
    dict(layer="Silver", table="products", column="brand",
         type="str",      nullable=True,  pk=False,
         description="Marca do produto.",
         validation="Pode ser nulo.",
         origin="Bronze", example="Acme"),
    dict(layer="Silver", table="products", column="created_ts",
         type="datetime", nullable=True,  pk=False,
         description="Data de cadastro do produto.",
         validation="NaT se inválido.",
         origin="Bronze", example="2024-03-18 00:00:00"),

    # ── GOLD ── fact_orders ──────────────────────────────────────────────────
    dict(layer="Gold", table="fact_orders", column="order_id",
         type="str",      nullable=False, pk=True,
         description="Chave do pedido. Grain: 1 linha por pedido.",
         validation="Único.",
         origin="Silver orders", example="O000001"),
    dict(layer="Gold", table="fact_orders", column="customer_id",
         type="str",      nullable=False, pk=False,
         description="FK para dim_customers.",
         validation="Não nulo.",
         origin="Silver orders", example="C00083"),
    dict(layer="Gold", table="fact_orders", column="order_date",
         type="datetime", nullable=True,  pk=False,
         description="Data do pedido (sem horário) — usada para agrupamentos diários.",
         validation="Derivada de order_ts.",
         origin="Silver orders", example="2025-12-01"),
    dict(layer="Gold", table="fact_orders", column="order_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp completo do pedido.",
         validation="NaT se ausente.",
         origin="Silver orders", example="2025-12-01 05:57:00"),
    dict(layer="Gold", table="fact_orders", column="gross_amount",
         type="float",    nullable=True,  pk=False,
         description="Receita bruta do pedido (total_amount da origem).",
         validation="float ≥ 0.",
         origin="Silver orders", example="48.82"),
    dict(layer="Gold", table="fact_orders", column="discount_total",
         type="float",    nullable=False, pk=False,
         description="Soma dos descontos de todos os itens do pedido.",
         validation="float ≥ 0; 0 se não houver itens.",
         origin="Silver order_items (agg)", example="105.83"),
    dict(layer="Gold", table="fact_orders", column="net_amount",
         type="float",    nullable=True,  pk=False,
         description="Receita líquida = soma de item_net_amount dos itens.",
         validation="Fallback para gross_amount se itens ausentes.",
         origin="Silver order_items (agg)", example="1010.04"),
    dict(layer="Gold", table="fact_orders", column="payment_method",
         type="str",      nullable=True,  pk=False,
         description="Forma de pagamento normalizada.",
         validation="Valores: pix, credit_card, debit_card.",
         origin="Silver orders", example="pix"),
    dict(layer="Gold", table="fact_orders", column="status_final",
         type="str",      nullable=True,  pk=False,
         description="Status final do pedido.",
         validation="Renomeado de 'status'.",
         origin="Silver orders", example="paid"),
    dict(layer="Gold", table="fact_orders", column="carrier",
         type="str",      nullable=True,  pk=False,
         description="Transportadora responsável pela entrega.",
         validation="Pode ser nulo (pedido não despachado).",
         origin="Silver shipments", example="CORREIOS"),
    dict(layer="Gold", table="fact_orders", column="shipping_cost",
         type="float",    nullable=True,  pk=False,
         description="Custo do frete em BRL.",
         validation="float ≥ 0.",
         origin="Silver shipments", example="39.57"),
    dict(layer="Gold", table="fact_orders", column="shipped_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp de despacho.",
         validation="NaT se não despachado.",
         origin="Silver shipments", example="2025-12-02 07:22:00"),
    dict(layer="Gold", table="fact_orders", column="delivered_ts",
         type="datetime", nullable=True,  pk=False,
         description="Timestamp de entrega.",
         validation="NaT se não entregue.",
         origin="Silver shipments", example="2025-12-07 00:08:00"),
    dict(layer="Gold", table="fact_orders", column="delivery_time_hours",
         type="float",    nullable=True,  pk=False,
         description="Tempo de entrega em horas: (delivered_ts − shipped_ts) / 3600.",
         validation="Nulo se shipped_ts ou delivered_ts ausentes.",
         origin="Derivada", example="112.77"),
    dict(layer="Gold", table="fact_orders", column="is_late",
         type="bool",     nullable=True,  pk=False,
         description="True se delivery_time_hours > 72h (regra configurável em settings.py).",
         validation="Nulo se delivery_time_hours for nulo.",
         origin="Derivada", example="True"),

    # ── GOLD ── fact_order_items ─────────────────────────────────────────────
    dict(layer="Gold", table="fact_order_items", column="order_id",
         type="str",   nullable=False, pk=True,
         description="FK para fact_orders.",
         validation="Não nulo.",
         origin="Silver order_items", example="O000001"),
    dict(layer="Gold", table="fact_order_items", column="product_id",
         type="str",   nullable=False, pk=True,
         description="FK para dim_products.",
         validation="Não nulo.",
         origin="Silver order_items", example="P00029"),
    dict(layer="Gold", table="fact_order_items", column="quantity",
         type="int",   nullable=False, pk=False,
         description="Quantidade de unidades compradas.",
         validation="int ≥ 0.",
         origin="Silver order_items", example="2"),
    dict(layer="Gold", table="fact_order_items", column="unit_price",
         type="float", nullable=True,  pk=False,
         description="Preço unitário em BRL.",
         validation="float ≥ 0.",
         origin="Silver order_items", example="567.16"),
    dict(layer="Gold", table="fact_order_items", column="discount_amount",
         type="float", nullable=False, pk=False,
         description="Desconto aplicado ao item.",
         validation="float ≥ 0.",
         origin="Silver order_items", example="21.74"),
    dict(layer="Gold", table="fact_order_items", column="item_net_amount",
         type="float", nullable=True,  pk=False,
         description="Receita líquida do item: (quantity × unit_price) − discount_amount.",
         validation="Pode ser negativo se desconto > preço.",
         origin="Silver order_items", example="1113.32"),

    # ── GOLD ── dim_customers ────────────────────────────────────────────────
    dict(layer="Gold", table="dim_customers", column="customer_id",
         type="str",      nullable=False, pk=True,
         description="Chave da dimensão cliente.",
         validation="Único.",
         origin="Silver customers", example="C00001"),
    dict(layer="Gold", table="dim_customers", column="state",
         type="str",      nullable=True,  pk=False,
         description="UF do cliente (2 chars maiúsculos).",
         validation="Padrão IBGE.",
         origin="Silver customers", example="SP"),
    dict(layer="Gold", table="dim_customers", column="city",
         type="str",      nullable=True,  pk=False,
         description="Cidade do cliente em Title Case.",
         validation="Texto normalizado.",
         origin="Silver customers", example="São Paulo"),
    dict(layer="Gold", table="dim_customers", column="created_ts",
         type="datetime", nullable=True,  pk=False,
         description="Data de cadastro do cliente.",
         validation="NaT se ausente.",
         origin="Silver customers", example="2023-04-13 00:00:00"),

    # ── GOLD ── dim_products ─────────────────────────────────────────────────
    dict(layer="Gold", table="dim_products", column="product_id",
         type="str",      nullable=False, pk=True,
         description="Chave da dimensão produto.",
         validation="Único.",
         origin="Silver products", example="P00001"),
    dict(layer="Gold", table="dim_products", column="category",
         type="str",      nullable=True,  pk=False,
         description="Categoria do produto.",
         validation="Texto normalizado.",
         origin="Silver products", example="Eletrônicos"),
    dict(layer="Gold", table="dim_products", column="brand",
         type="str",      nullable=True,  pk=False,
         description="Marca do produto.",
         validation="Pode ser nulo.",
         origin="Silver products", example="Acme"),
    dict(layer="Gold", table="dim_products", column="created_ts",
         type="datetime", nullable=True,  pk=False,
         description="Data de cadastro do produto.",
         validation="NaT se ausente.",
         origin="Silver products", example="2024-03-18 00:00:00"),
]

# Metadata columns present in Bronze and Silver (_ingest_date, _source_file, _load_ts)
_META_COLUMNS = [
    ("_ingest_date", "str",   False, "Data de ingestão no pipeline (YYYY-MM-DD)."),
    ("_source_file", "str",   False, "Nome do arquivo de origem."),
    ("_load_ts",     "str",   False, "Timestamp ISO 8601 em que a linha foi carregada."),
]

_BRONZE_SILVER_TABLES = [
    "orders", "order_items", "shipments", "customers", "products",
]

for _layer in ("Bronze", "Silver"):
    for _tbl in _BRONZE_SILVER_TABLES:
        for _col, _type, _nullable, _desc in _META_COLUMNS:
            CATALOG.append(dict(
                layer=_layer, table=_tbl, column=_col,
                type=_type, nullable=_nullable, pk=False,
                description=_desc,
                validation="Adicionado automaticamente pelo BronzeIngestor.",
                origin="Pipeline", example="2025-12-01",
            ))


# ─────────────────────────────────────────────────────────────────────────────
# TABLE OVERVIEW DEFINITION
# ─────────────────────────────────────────────────────────────────────────────

TABLE_OVERVIEW: list[dict] = [
    # Bronze
    dict(layer="Bronze", table="orders",      grain="1 linha por order_id",
         pk="order_id",         description="Pedidos ingeridos da fonte sem transformação.",
         destination="data/bronze/ingest_date=*/orders.csv"),
    dict(layer="Bronze", table="order_items", grain="1 linha por (order_id, product_id)",
         pk="order_id, product_id", description="Itens de cada pedido ingeridos da fonte.",
         destination="data/bronze/ingest_date=*/order_items.csv"),
    dict(layer="Bronze", table="shipments",   grain="1 linha por order_id",
         pk="order_id",         description="Dados de envio e entrega por pedido.",
         destination="data/bronze/ingest_date=*/shipments.csv"),
    dict(layer="Bronze", table="customers",   grain="1 linha por customer_id",
         pk="customer_id",      description="Cadastro de clientes.",
         destination="data/bronze/ingest_date=*/customers.csv"),
    dict(layer="Bronze", table="products",    grain="1 linha por product_id",
         pk="product_id",       description="Catálogo de produtos.",
         destination="data/bronze/ingest_date=*/products.csv"),
    # Silver
    dict(layer="Silver", table="orders",      grain="1 linha por order_id",
         pk="order_id",         description="Pedidos limpos: tipos corretos, deduplicados, status/currency normalizados.",
         destination="data/silver/orders.csv"),
    dict(layer="Silver", table="order_items", grain="1 linha por (order_id, product_id)",
         pk="order_id, product_id", description="Itens com item_net_amount calculado.",
         destination="data/silver/order_items.csv"),
    dict(layer="Silver", table="shipments",   grain="1 linha por order_id",
         pk="order_id",         description="Shipments com timestamps parseados e shipping_cost corrigido.",
         destination="data/silver/shipments.csv"),
    dict(layer="Silver", table="customers",   grain="1 linha por customer_id",
         pk="customer_id",      description="Clientes deduplicados com UF/cidade normalizados.",
         destination="data/silver/customers.csv"),
    dict(layer="Silver", table="products",    grain="1 linha por product_id",
         pk="product_id",       description="Produtos deduplicados.",
         destination="data/silver/products.csv"),
    # Gold
    dict(layer="Gold",   table="fact_orders",      grain="1 linha por order_id",
         pk="order_id",         description="Fato principal: receita bruta/líquida, descontos, KPIs de entrega (is_late, delivery_time_hours).",
         destination="data/gold/fact_orders.csv"),
    dict(layer="Gold",   table="fact_order_items",  grain="1 linha por (order_id, product_id)",
         pk="order_id, product_id", description="Fato de itens: receita líquida por item.",
         destination="data/gold/fact_order_items.csv"),
    dict(layer="Gold",   table="dim_customers",     grain="1 linha por customer_id",
         pk="customer_id",      description="Dimensão cliente com UF e cidade para segmentação geográfica.",
         destination="data/gold/dim_customers.csv"),
    dict(layer="Gold",   table="dim_products",      grain="1 linha por product_id",
         pk="product_id",       description="Dimensão produto com categoria e marca.",
         destination="data/gold/dim_products.csv"),
]


# ─────────────────────────────────────────────────────────────────────────────
# QUALITY PROFILE  (reads actual Silver/Gold files if available)
# ─────────────────────────────────────────────────────────────────────────────

def _profile_table(layer: str, table: str) -> dict:
    """Return row count and per-column null rate for an existing Silver/Gold CSV."""
    if layer == "Silver":
        path = SILVER_DIR / f"{table}.csv"
    elif layer == "Gold":
        path = GOLD_DIR / f"{table}.csv"
    else:
        return {}

    if not path.exists():
        return {}

    df = pd.read_csv(path)
    profile = {"rows": len(df)}
    for col in df.columns:
        null_rate = df[col].isna().mean()
        profile[col] = f"{null_rate:.1%} nulos"
    return profile


def build_quality_sheet() -> pd.DataFrame:
    rows = []
    silver_tables = ["orders", "order_items", "shipments", "customers", "products"]
    gold_tables   = ["fact_orders", "fact_order_items", "dim_customers", "dim_products"]

    for layer, tables in [("Silver", silver_tables), ("Gold", gold_tables)]:
        for table in tables:
            profile = _profile_table(layer, table)
            if not profile:
                rows.append({"Layer": layer, "Tabela": table,
                             "Linhas": "—", "Coluna": "—", "% Nulos": "Arquivo não encontrado"})
                continue
            row_count = profile.pop("rows")
            for col, null_info in profile.items():
                rows.append({
                    "Layer": layer,
                    "Tabela": table,
                    "Linhas": row_count,
                    "Coluna": col,
                    "% Nulos": null_info,
                })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT WITH FORMATTING
# ─────────────────────────────────────────────────────────────────────────────

# Layer colour palette (header background)
_LAYER_COLORS = {
    "Bronze": "C65911",   # dark orange
    "Silver": "44546A",   # slate
    "Gold":   "7F6000",   # dark gold
}

_LAYER_LIGHT = {
    "Bronze": "FCE4D6",
    "Silver": "D6DCE4",
    "Gold":   "FFF2CC",
}


def _apply_header_style(ws, color_hex: str) -> None:
    """Bold + colored background on the first row of a worksheet."""
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill("solid", fgColor=color_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws) -> None:
    """Set column widths based on max content length."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def generate(output_path: Path | None = None) -> Path:
    """Build the data dictionary Excel file and return its path."""
    try:
        from openpyxl.styles import PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — run: pip install openpyxl")
        raise

    output_path = output_path or Path(__file__).parent / "data_dictionary.xlsx"

    # ── Sheet 1: Visão Geral ──────────────────────────────────────────────────
    overview_df = pd.DataFrame(TABLE_OVERVIEW).rename(columns={
        "layer": "Layer", "table": "Tabela", "grain": "Granularidade",
        "pk": "Chave Primária", "description": "Descrição", "destination": "Caminho",
    })

    # ── Sheet 2: Colunas ──────────────────────────────────────────────────────
    columns_df = pd.DataFrame(CATALOG).rename(columns={
        "layer": "Layer", "table": "Tabela", "column": "Coluna",
        "type": "Tipo", "nullable": "Nullable", "pk": "PK",
        "description": "Descrição", "validation": "Validação / Regra",
        "origin": "Origem", "example": "Exemplo",
    })
    columns_df["Nullable"] = columns_df["Nullable"].map({True: "Sim", False: "Não"})
    columns_df["PK"]       = columns_df["PK"].map({True: "✓", False: ""})

    # ── Sheet 3: Qualidade ────────────────────────────────────────────────────
    quality_df = build_quality_sheet()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Visão Geral", index=False)
        columns_df.to_excel(writer,  sheet_name="Colunas",     index=False)
        quality_df.to_excel(writer,  sheet_name="Qualidade",   index=False)

        wb = writer.book

        # Style each sheet
        for sheet_name, color in [
            ("Visão Geral", "1F4E79"),
            ("Colunas",     "1F4E79"),
            ("Qualidade",   "375623"),
        ]:
            ws = wb[sheet_name]
            _apply_header_style(ws, color)
            _auto_width(ws)

            # Zebra rows + layer colour tint for Colunas sheet
            if sheet_name == "Colunas":
                layer_col_idx = columns_df.columns.get_loc("Layer") + 1
                for row in ws.iter_rows(min_row=2):
                    layer_val = row[layer_col_idx - 1].value or ""
                    light = _LAYER_LIGHT.get(layer_val, "FFFFFF")
                    fill  = PatternFill("solid", fgColor=light)
                    for cell in row:
                        cell.fill = fill
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

    print(f"Data dictionary generated: {output_path}")
    return output_path


if __name__ == "__main__":
    generate()
